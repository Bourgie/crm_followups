import os
import io
import hashlib
from datetime import datetime, timedelta
from pathlib import Path
from urllib.parse import urlencode

import requests
from dotenv import load_dotenv
from openpyxl import Workbook

from fastapi import FastAPI, UploadFile, File, Form, Request, HTTPException
from fastapi.responses import HTMLResponse, RedirectResponse, JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from starlette.middleware.sessions import SessionMiddleware

from .pdf_parser import parse_budget_pdf
from .calendar_service import (
    create_followup_events,
    delete_events,
    create_postventa_event,
)
from .google_oauth import (
    get_auth_url,
    exchange_code_for_creds,
    token_path_for_email,
    load_creds_for_email,
    save_creds_for_email,
)
from .db import (
    init_db,
    find_existing,
    insert_quote,
    list_quotes,
    get_quote_detail,
    update_notes,
    clear_events,
    get_event_ids,
    # postventas
    insert_postventa,
    list_postventas,
    get_postventa_detail,
    update_postventa_status,
    clear_postventa_event,
    # admin
    list_vendors,
    list_admin_items,
    # KPIs
    get_kpis,
    list_vendor_kpis_month,
)

load_dotenv()

APP_DIR = Path(__file__).resolve().parent.parent
UPLOADS_DIR = APP_DIR / "uploads"
UPLOADS_DIR.mkdir(exist_ok=True)

TEMPLATES_DIR = APP_DIR / "templates"
STATIC_DIR = APP_DIR / "static"

app = FastAPI(title="CRM Followups MVP")

# -----------------------------
# Middleware: Sessions
# -----------------------------
app.add_middleware(
    SessionMiddleware,
    secret_key=os.getenv("SESSION_SECRET", "cambia-esto-por-una-clave-larga"),
    same_site="lax",
    https_only=os.getenv("HTTPS_ONLY", "0") == "1",  # local 0; prod 1 con HTTPS
)

# -----------------------------
# Static + Templates
# -----------------------------
app.mount("/static", StaticFiles(directory=str(STATIC_DIR)), name="static")
templates = Jinja2Templates(directory=str(TEMPLATES_DIR))

# -----------------------------
# DB init
# -----------------------------
init_db()


# -----------------------------
# Helpers
# -----------------------------
def current_email(request: Request) -> str | None:
    return request.session.get("vendor_email")


def get_google_user_email(creds) -> str:
    r = requests.get(
        "https://www.googleapis.com/oauth2/v2/userinfo",
        headers={"Authorization": f"Bearer {creds.token}"},
        timeout=15,
    )
    r.raise_for_status()
    return r.json()["email"]


def redirect_with_msg(url: str, msg: str, msg_type: str = "success") -> RedirectResponse:
    # Maneja tildes, símbolos, etc. de forma correcta
    qs = urlencode({"msg": msg, "msg_type": msg_type})
    return RedirectResponse(f"{url}?{qs}", status_code=303)


def admin_emails() -> set[str]:
    raw = os.getenv("ADMIN_EMAILS", "")
    return {e.strip().lower() for e in raw.split(",") if e.strip()}


def is_admin_email(email: str) -> bool:
    return (email or "").strip().lower() in admin_emails()


def require_admin(request: Request) -> str:
    email = current_email(request)
    if not email:
        raise HTTPException(status_code=401, detail="No logueado")
    if not is_admin_email(email):
        raise HTTPException(status_code=403, detail="No autorizado (admin)")
    return email


def save_uploaded_pdf(pdf: UploadFile, content: bytes) -> Path:
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_name = (pdf.filename or "archivo.pdf").replace("/", "_").replace("\\", "_")
    out_path = UPLOADS_DIR / f"{ts}_{safe_name}"
    out_path.write_bytes(content)
    return out_path


# -----------------------------
# Auth
# -----------------------------
@app.get("/", response_class=HTMLResponse)
def root(request: Request):
    return RedirectResponse("/ui", status_code=303) if current_email(request) else RedirectResponse("/login", status_code=303)


@app.get("/login")
def login():
    auth_url = get_auth_url(vendor_id="login")
    return RedirectResponse(auth_url, status_code=303)


@app.get("/logout")
def logout(request: Request):
    request.session.clear()
    return RedirectResponse("/login", status_code=303)


@app.get("/auth/callback")
def auth_callback(request: Request, code: str, state: str):
    creds = exchange_code_for_creds(code=code, vendor_id=state)
    email = get_google_user_email(creds)

    save_creds_for_email(email, creds)

    request.session["vendor_email"] = email
    return RedirectResponse("/ui", status_code=303)


# -----------------------------
# UI (vendedor)
# -----------------------------
@app.get("/ui", response_class=HTMLResponse)
def ui_home(request: Request, msg: str = "", msg_type: str = ""):
    email = current_email(request)
    if not email:
        return RedirectResponse("/login", status_code=303)

    is_connected = token_path_for_email(email).exists()
    kpis = get_kpis(vendor_id=email, older_than_days=7)

    return templates.TemplateResponse(
        "ui_home.html",
        {
            "request": request,
            "vendor_email": email,
            "is_connected": is_connected,
            "quotes": list_quotes(email),
            "postventas": list_postventas(email),
            "kpis": kpis,
            "is_admin": is_admin_email(email),
            "msg": msg,
            "msg_type": msg_type,
        },
    )


@app.post("/ui/upload")
async def ui_upload(request: Request, pdf: UploadFile = File(...)):
    email = current_email(request)
    if not email:
        return RedirectResponse("/login", status_code=303)

    if not (pdf.filename or "").lower().endswith(".pdf"):
        return redirect_with_msg("/ui", "Archivo no es PDF", "error")

    content = await pdf.read()
    pdf_sha256 = hashlib.sha256(content).hexdigest()

    out_path = save_uploaded_pdf(pdf, content)

    creds = load_creds_for_email(email)
    if creds is None:
        request.session.clear()
        return RedirectResponse("/login", status_code=303)

    extracted = parse_budget_pdf(str(out_path))
    quote_number = extracted.get("quote_number", "S/N")

    existing = find_existing(vendor_id=email, quote_number=quote_number, pdf_sha256=pdf_sha256)
    if existing:
        return redirect_with_msg("/ui", "Duplicado bloqueado (ya existía).", "success")

    events_created = create_followup_events(creds=creds, quote_data=extracted)

    insert_quote(
        vendor_id=email,
        quote_number=quote_number,
        pdf_sha256=pdf_sha256,
        extracted=extracted,
        events_created=events_created,
    )

    return redirect_with_msg("/ui", "Listo: recordatorios creados (48h y 72h).", "success")


@app.get("/ui/quote", response_class=HTMLResponse)
def ui_quote(request: Request, quote_number: str, msg: str = "", msg_type: str = ""):
    email = current_email(request)
    if not email:
        return RedirectResponse("/login", status_code=303)

    detail = get_quote_detail(vendor_id=email, quote_number=quote_number)
    if not detail:
        return redirect_with_msg("/ui", "Cotización no encontrada", "error")

    return templates.TemplateResponse(
        "ui_quote.html",
        {
            "request": request,
            "vendor_email": email,
            "is_admin": is_admin_email(email),
            "detail": detail,
            "msg": msg,
            "msg_type": msg_type,
        },
    )


@app.post("/ui/quote/save")
def ui_quote_save(
    request: Request,
    quote_number: str = Form(...),
    summary: str = Form(""),
    notes: str = Form(""),
    status: str = Form("pendiente"),
):
    email = current_email(request)
    if not email:
        return RedirectResponse("/login", status_code=303)

    update_notes(vendor_id=email, quote_number=quote_number, summary=summary, notes=notes, status=status)

    new_s = (status or "").strip().lower()

    # Si pasa a cerrada/perdida => cancelar eventos (si existían) + limpiar events_json
    if new_s in ("cerrada", "perdida"):
        event_ids = get_event_ids(vendor_id=email, quote_number=quote_number)
        print(f"[AUTO CANCEL] quote={quote_number} new={new_s} event_ids={event_ids}")

        if not event_ids:
            clear_events(vendor_id=email, quote_number=quote_number)
            return redirect_with_msg(
                f"/ui/quote?quote_number={quote_number}",
                "Guardado OK (sin eventos para cancelar)",
                "success",
            )

        creds = load_creds_for_email(email)
        if creds is None:
            return redirect_with_msg(
                f"/ui/quote?quote_number={quote_number}",
                "No hay credenciales para borrar eventos",
                "error",
            )

        result = delete_events(creds=creds, event_ids=event_ids, calendar_id="primary")
        print(f"[AUTO CANCEL RESULT] {result}")

        if result.get("failed"):
            return redirect_with_msg(
                f"/ui/quote?quote_number={quote_number}",
                "Error borrando eventos (ver consola)",
                "error",
            )

        clear_events(vendor_id=email, quote_number=quote_number)
        print(f"[AUTO CANCEL] events_json limpiado para {quote_number}")

    return redirect_with_msg(
        f"/ui/quote?quote_number={quote_number}",
        "Guardado OK",
        "success",
    )


@app.post("/ui/quote/cancel")
def ui_quote_cancel(request: Request, quote_number: str = Form(...)):
    email = current_email(request)
    if not email:
        return RedirectResponse("/login", status_code=303)

    event_ids = get_event_ids(vendor_id=email, quote_number=quote_number)
    print(f"[MANUAL CANCEL] quote={quote_number} event_ids={event_ids}")

    if not event_ids:
        clear_events(vendor_id=email, quote_number=quote_number)
        return redirect_with_msg(
            f"/ui/quote?quote_number={quote_number}",
            "No había recordatorios para cancelar",
            "success",
        )

    creds = load_creds_for_email(email)
    if creds is None:
        return redirect_with_msg(
            f"/ui/quote?quote_number={quote_number}",
            "No hay credenciales para cancelar",
            "error",
        )

    result = delete_events(creds=creds, event_ids=event_ids, calendar_id="primary")
    print(f"[MANUAL CANCEL RESULT] {result}")

    if result.get("failed"):
        return redirect_with_msg(
            f"/ui/quote?quote_number={quote_number}",
            "Error cancelando (ver consola)",
            "error",
        )

    clear_events(vendor_id=email, quote_number=quote_number)
    return redirect_with_msg(
        f"/ui/quote?quote_number={quote_number}",
        "Recordatorios cancelados",
        "success",
    )


# -----------------------------
# Postventa (manual + desde cotización cerrada)
# -----------------------------
@app.get("/ui/postventa/new", response_class=HTMLResponse)
def ui_postventa_new(request: Request, msg: str = "", msg_type: str = ""):
    email = current_email(request)
    if not email:
        return RedirectResponse("/login", status_code=303)

    default_day = (datetime.now() + timedelta(days=7)).strftime("%Y-%m-%d")

    return templates.TemplateResponse(
        "ui_postventa_new.html",
        {
            "request": request,
            "vendor_email": email,
            "default_postventa_date": default_day,
            "msg": msg,
            "msg_type": msg_type,
            "is_admin": is_admin_email(email),
        },
    )


@app.post("/ui/postventa/create")
def ui_postventa_create(
    request: Request,
    client_name: str = Form(...),
    phone: str = Form(""),
    sale_date: str = Form(""),
    postventa_date: str = Form(...),
    type: str = Form("postventa"),
    notes: str = Form(""),
):
    email = current_email(request)
    if not email:
        return RedirectResponse("/login", status_code=303)

    creds = load_creds_for_email(email)
    if creds is None:
        request.session.clear()
        return RedirectResponse("/login", status_code=303)

    data = {
        "client_name": client_name.strip(),
        "phone": phone.strip(),
        "sale_date": sale_date.strip(),
        "postventa_date": postventa_date.strip(),
        "type": type.strip(),
        "notes": notes.strip(),
    }

    event = create_postventa_event(creds=creds, data=data, calendar_id="primary")

    new_id = insert_postventa(
        vendor_id=email,
        client_name=data["client_name"],
        phone=data["phone"],
        sale_date=data["sale_date"],
        postventa_date=data["postventa_date"],
        type_=data["type"],
        notes=data["notes"],
        event=event,
    )

    return redirect_with_msg(
        f"/ui/postventa?postventa_id={new_id}",
        "Postventa creada",
        "success",
    )


@app.get("/ui/postventa", response_class=HTMLResponse)
def ui_postventa_detail(request: Request, postventa_id: int, msg: str = "", msg_type: str = ""):
    email = current_email(request)
    if not email:
        return RedirectResponse("/login", status_code=303)

    pv = get_postventa_detail(vendor_id=email, postventa_id=postventa_id)
    if not pv:
        return redirect_with_msg("/ui", "Postventa no encontrada", "error")

    return templates.TemplateResponse(
        "ui_postventa_detail.html",
        {
            "request": request,
            "vendor_email": email,
            "pv": pv,
            "msg": msg,
            "msg_type": msg_type,
            "is_admin": is_admin_email(email),
        },
    )


@app.post("/ui/postventa/done")
def ui_postventa_done(request: Request, postventa_id: int = Form(...)):
    email = current_email(request)
    if not email:
        return RedirectResponse("/login", status_code=303)

    update_postventa_status(vendor_id=email, postventa_id=postventa_id, status="realizada")

    return redirect_with_msg(
        f"/ui/postventa?postventa_id={postventa_id}",
        "Marcada como realizada",
        "success",
    )


@app.post("/ui/postventa/cancel")
def ui_postventa_cancel(request: Request, postventa_id: int = Form(...)):
    email = current_email(request)
    if not email:
        return RedirectResponse("/login", status_code=303)

    pv = get_postventa_detail(vendor_id=email, postventa_id=postventa_id)
    if not pv:
        return redirect_with_msg("/ui", "Postventa no encontrada", "error")

    event_id = pv.get("event_id")
    if event_id:
        creds = load_creds_for_email(email)
        if creds is None:
            return redirect_with_msg(
                f"/ui/postventa?postventa_id={postventa_id}",
                "No hay credenciales",
                "error",
            )

        result = delete_events(creds=creds, event_ids=[event_id], calendar_id="primary")
        print("[POSTVENTA CANCEL RESULT]", result)

        if result.get("failed"):
            return redirect_with_msg(
                f"/ui/postventa?postventa_id={postventa_id}",
                "Error borrando evento (ver consola)",
                "error",
            )

    update_postventa_status(vendor_id=email, postventa_id=postventa_id, status="cancelada")
    clear_postventa_event(vendor_id=email, postventa_id=postventa_id)

    return redirect_with_msg(
        f"/ui/postventa?postventa_id={postventa_id}",
        "Postventa cancelada",
        "success",
    )


@app.post("/ui/quote/postventa")
def ui_quote_create_postventa(request: Request, quote_number: str = Form(...)):
    email = current_email(request)
    if not email:
        return RedirectResponse("/login", status_code=303)

    detail = get_quote_detail(vendor_id=email, quote_number=quote_number)
    if not detail:
        return redirect_with_msg("/ui", "Cotización no encontrada", "error")

    status = (detail.get("status") or "").strip().lower()
    if status != "cerrada":
        return redirect_with_msg(
            f"/ui/quote?quote_number={quote_number}",
            "Solo disponible si está cerrada",
            "error",
        )

    extracted = detail.get("extracted") or {}
    client_name = (extracted.get("client_name") or "").strip() or "Cliente"
    sale_date = (extracted.get("issue_date") or "").strip()
    postventa_date = (datetime.now() + timedelta(days=7)).strftime("%Y-%m-%d")

    creds = load_creds_for_email(email)
    if creds is None:
        request.session.clear()
        return RedirectResponse("/login", status_code=303)

    data = {
        "client_name": client_name,
        "phone": "",
        "sale_date": sale_date,
        "postventa_date": postventa_date,
        "type": "postventa",
        "notes": f"Postventa creada desde cotización {quote_number}.",
    }

    event = create_postventa_event(creds=creds, data=data, calendar_id="primary")

    new_id = insert_postventa(
        vendor_id=email,
        client_name=data["client_name"],
        phone=data["phone"],
        sale_date=data["sale_date"],
        postventa_date=data["postventa_date"],
        type_=data["type"],
        notes=data["notes"],
        event=event,
    )

    return redirect_with_msg(
        f"/ui/postventa?postventa_id={new_id}",
        "Postventa creada +7 días",
        "success",
    )


# -----------------------------
# Admin (encargado)
# -----------------------------
@app.get("/admin", response_class=HTMLResponse)
def admin_home(
    request: Request,
    vendor_id: str = "",
    status: str = "",
    kind: str = "all",
    date_from: str = "",
    date_to: str = "",
    msg: str = "",
    msg_type: str = "",
):
    admin_email = require_admin(request)

    vendors = list_vendors()

    items = list_admin_items(
        vendor_id=vendor_id.strip() or None,
        status=status.strip() or None,
        date_from=date_from.strip() or None,
        date_to=date_to.strip() or None,
        kind=(kind.strip() or "all"),
        limit=500,
    )

    # KPIs globales del mes (todo)
    kpis_global = get_kpis(vendor_id=None, older_than_days=7)

    # KPIs del mes del vendedor seleccionado (si aplica)
    kpis_vendor = get_kpis(vendor_id=vendor_id.strip() or None, older_than_days=7) if vendor_id.strip() else None

    # Ranking por vendedor del mes
    ranking = list_vendor_kpis_month()

    return templates.TemplateResponse(
        "admin_home.html",
        {
            "request": request,
            "admin_email": admin_email,
            "vendors": vendors,
            "items": items,
            "kpis_global": kpis_global,
            "kpis_vendor": kpis_vendor,
            "ranking": ranking,
            "filters": {
                "vendor_id": vendor_id,
                "status": status,
                "kind": kind,
                "date_from": date_from,
                "date_to": date_to,
            },
            "msg": msg,
            "msg_type": msg_type,
        },
    )


@app.get("/admin/export.xlsx")
def admin_export_excel(
    request: Request,
    vendor_id: str = "",
    status: str = "",
    kind: str = "all",
    date_from: str = "",
    date_to: str = "",
):
    require_admin(request)

    vendor_filter = vendor_id.strip() or None
    df = date_from.strip() or None
    dt = date_to.strip() or None
    st = status.strip() or None
    kd = kind.strip() or "all"

    items = list_admin_items(
        vendor_id=vendor_filter,
        status=st,
        date_from=df,
        date_to=dt,
        kind=kd,
        limit=5000,
    )

    wb = Workbook()

    # --- Sheet KPIs ---
    ws0 = wb.active
    ws0.title = "KPIs"

    g = get_kpis(vendor_id=None, older_than_days=7)
    v = get_kpis(vendor_id=vendor_filter, older_than_days=7) if vendor_filter else None

    ws0.append(["KPIs", "Valor"])
    ws0.append(["Mes actual", f"{g['month']['from']} → {g['month']['to']}"])
    ws0.append(["Cotizaciones (mes)", g["month"]["quotes"]["total"]])
    ws0.append(["Cerradas (mes)", g["month"]["quotes"]["cerrada"]])
    ws0.append(["Perdidas (mes)", g["month"]["quotes"]["perdida"]])
    ws0.append(["% cierre (mes)", g["month"]["close_rate"] if g["month"]["close_rate"] is not None else ""])
    ws0.append(["Postventas (mes)", g["month"]["postventas"]["total"]])
    ws0.append(["Alertas: abiertas +7d", g["alerts"]["old_open_quotes"]])

    if v:
        ws0.append([])
        ws0.append(["KPIs vendedor", vendor_filter])
        ws0.append(["Cotizaciones (mes)", v["month"]["quotes"]["total"]])
        ws0.append(["Cerradas (mes)", v["month"]["quotes"]["cerrada"]])
        ws0.append(["Perdidas (mes)", v["month"]["quotes"]["perdida"]])
        ws0.append(["% cierre (mes)", v["month"]["close_rate"] if v["month"]["close_rate"] is not None else ""])
        ws0.append(["Postventas (mes)", v["month"]["postventas"]["total"]])

    # --- Sheet Items (filtros actuales) ---
    ws1 = wb.create_sheet("Items")
    ws1.append(["date", "kind", "vendor", "client_name", "ref", "status", "total", "summary", "created_at"])
    for it in items:
        ws1.append([
            it.get("date", ""),
            it.get("kind", ""),
            it.get("vendor_id", ""),
            it.get("client_name", ""),
            it.get("ref", ""),
            it.get("status", ""),
            it.get("total", ""),
            it.get("summary", ""),
            it.get("created_at", ""),
        ])

    # --- Sheet Ranking ---
    ws2 = wb.create_sheet("Ranking vendedores (mes)")
    ws2.append(["vendor", "quotes_total", "cerradas", "perdidas", "pendientes", "%cierre", "postventas_total"])
    for r in list_vendor_kpis_month():
        ws2.append([
            r["vendor_id"],
            r["quotes_total"],
            r["quotes_cerrada"],
            r["quotes_perdida"],
            r["quotes_pendiente"],
            r["close_rate"] if r["close_rate"] is not None else "",
            r["postventas_total"],
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = "reporte_admin.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# -----------------------------
# API (Swagger) opcional
# -----------------------------
@app.post("/upload")
async def api_upload_pdf(request: Request, pdf: UploadFile = File(...)):
    email = current_email(request)
    if not email:
        return JSONResponse({"error": "No logueado. Abrí /login y autorizá."}, status_code=401)

    if not (pdf.filename or "").lower().endswith(".pdf"):
        return JSONResponse({"error": "Subí un archivo .pdf"}, status_code=400)

    content = await pdf.read()
    pdf_sha256 = hashlib.sha256(content).hexdigest()
    out_path = save_uploaded_pdf(pdf, content)

    creds = load_creds_for_email(email)
    if creds is None:
        return JSONResponse({"error": "No hay credenciales. Iniciá sesión en /login."}, status_code=400)

    extracted = parse_budget_pdf(str(out_path))
    quote_number = extracted.get("quote_number", "S/N")

    existing = find_existing(vendor_id=email, quote_number=quote_number, pdf_sha256=pdf_sha256)
    if existing:
        return {
            "status": "DUPLICATE_BLOCKED",
            "message": "Ya existía esta cotización para este vendedor. No se crearon eventos nuevos.",
            "existing_record": existing,
        }

    events_created = create_followup_events(creds=creds, quote_data=extracted)

    insert_quote(
        vendor_id=email,
        quote_number=quote_number,
        pdf_sha256=pdf_sha256,
        extracted=extracted,
        events_created=events_created,
    )

    return {
        "status": "OK",
        "vendor_email": email,
        "saved_pdf": str(out_path),
        "pdf_sha256": pdf_sha256,
        "extracted": extracted,
        "events_created": events_created,
    }
